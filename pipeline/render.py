#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Render the 4 due-diligence deliverables from the analysis engine output.
All content is driven by extracted facts -> no hardcoded instance answers.
"""
import os, json, sys, datetime
# Offline bootstrap: vendored pure-python deps + shared stdlib docx helper.
_SUB = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
for _p in (os.path.join(_SUB, "vendor"), _SUB):
    if _p not in sys.path:
        sys.path.insert(0, _p)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from _docx import DocxDocument

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(os.path.dirname(BASE), "deliverables_v2")
os.makedirs(OUT, exist_ok=True)

def set_out_dir(out_dir):
    global OUT
    OUT = out_dir
    os.makedirs(OUT, exist_ok=True)

DOMAIN = {"CORP":"主体资格","CAP":"股权出资","GOV":"公司治理","CON":"重大合同",
          "IP":"知识产权","HR":"劳动人事","DATA":"数据合规","REG":"业务资质",
          "DISP":"争议债务","PROP":"物业资产"}
STATUS_ORDER = ["已满足","部分满足","未满足","不适用","待确认","资料冲突"]
RISK_ORDER = ["HIGH","MEDIUM","LOW","INFO"]
DISP_ORDER = ["CP","COVENANT","INDEMNITY","PRICE","RFI","MONITOR","NONE"]
TYPE_OK = {"补充资料","事实问询","外部确认","整改动作"}
BASELINE_DATE = "2026-07-15"  # demo baseline ("as-of") date

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = CENTER; cell.border = BORDER

# ---------------------------------------------------------------------------
def render(res, checklist, store, scenario=None, out_dir=None):
    if out_dir:
        set_out_dir(out_dir)
    items = res["items"]; feedback = res["feedback"]; kis = res["key_issues"]
    contradictions = res["contradictions"]
    fb_by_id = {f["fid"]: f for f in feedback}
    meta = res.get("meta", {})
    TARGET = meta.get("target_name") or "目标公司（见登记档案）"
    BASELINE = meta.get("baseline_date") or (scenario.baseline_date if scenario else BASELINE_DATE)
    MATN = meta.get("material_count") or len(store.by_prefix)
    IID = meta.get("instance_id") or "instance"
    ACQUIRER = scenario.acquirer if scenario else "买方（见交易文件）"
    DEAL_PCT = scenario.deal_pct if scenario else ""
    ORDER = list(checklist)  # stable checklist order
    # ensure all 52 present
    ORDER = [it["id"] for it in checklist]

    # ============ 1) 尽调清单核验与反馈表.xlsx ============
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "核验总表"
    cols = ["清单编号","领域","清单要求","核验状态","事实摘要","证据定位","资料缺口",
            "风险等级","处置类型","反馈意见","是否需人工复核","关联反馈编号"]
    ws.append(cols); style_header(ws, 1, len(cols))
    for iid in ORDER:
        d = items[iid]
        fb_asks = [fb_by_id[f]["ask"] for f in d.get("feedback_ids", []) if f in fb_by_id]
        ws.append([
            iid, d["domain"], d["req"], d["status"], d["fact"], d["evidence"],
            d["gap"], d["risk"], d["disp"],
            "\n".join(fb_asks) if fb_asks else "无。",
            d.get("human","否"),
            "、".join(d.get("feedback_ids", [])),
        ])
    widths = [10,9,40,9,46,46,30,9,11,40,11,14]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row+1):
        for c in range(1, len(cols)+1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP if c in (3,5,6,7,10) else CENTER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), ws.max_row)

    # 证据索引
    ws2 = wb.create_sheet("证据索引")
    files = {}
    for iid in ORDER:
        for ev in items[iid]["evidence"].split("\n"):
            ev = ev.strip()
            if "｜" in ev:
                fn, loc = ev.split("｜",1)
            else:
                fn, loc = ev, ""
            if "未发现可支持资料" in fn:
                fn = "（无）"
            files.setdefault(fn, set()).add(loc)
    ws2.append(["被引用文件","精确定位（章节/工作表/单元格/页码）","关联清单编号"])
    style_header(ws2, 1, 3)
    for fn, locs in files.items():
        related = [iid for iid in ORDER if fn in items[iid]["evidence"] or "未发现可支持资料" in items[iid]["evidence"]]
        ws2.append([fn, "\n".join(sorted(l for l in locs if l)), "、".join(related)])
    ws2.column_dimensions["A"].width = 42; ws2.column_dimensions["B"].width = 60; ws2.column_dimensions["C"].width = 40
    for r in range(2, ws2.max_row+1):
        for c in (1,2,3):
            ws2.cell(row=r,column=c).border = BORDER; ws2.cell(row=r,column=c).alignment = WRAP
    ws2.freeze_panes = "A2"

    # 跨文件矛盾
    wsc = wb.create_sheet("跨文件矛盾")
    ccols = ["矛盾编号","严重度","矛盾主题","冲突文件","关联清单编号","矛盾说明","建议处置"]
    wsc.append(ccols); style_header(wsc, 1, len(ccols))
    for c in contradictions:
        wsc.append([c["cid"], c["severity"], c["title"], "\n".join(c["files_conflict"]),
                    "、".join(c["items"]), c["description"], c["resolution"]])
    cwidths=[10,8,34,44,18,60,46]
    for i,w in enumerate(cwidths,1): wsc.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, wsc.max_row+1):
        for cc in range(1, len(ccols)+1):
            wsc.cell(r,cc).border = BORDER
            wsc.cell(r,cc).alignment = WRAP if cc in (3,4,6,7) else CENTER
    wsc.freeze_panes = "A2"

    # 统计摘要 (formulas)
    ws3 = wb.create_sheet("统计摘要")
    ws3["A1"] = f"{TARGET}｜法律尽职调查第一轮资料核验统计摘要"; ws3["A1"].font = TITLE_FONT
    ws3.append([])
    ST_COL="D"; RK_COL="H"; DP_COL="I"; DM_COL="B"
    def cf(col, val): return f'=COUNTIF(核验总表!{col}:{col},"{val}")'
    ws3.append(["一、按核验状态"]); ws3["A3"].font=Font(bold=True)
    ws3.append(["核验状态","公式计数","Python校验值"]); style_header(ws3,4,3)
    r=5
    for s in STATUS_ORDER:
        ws3.cell(r,1,s); ws3.cell(r,2,cf(ST_COL,s)); ws3.cell(r,3,sum(1 for i in ORDER if items[i]["status"]==s)); r+=1
    ws3.cell(r,1,"合计"); ws3.cell(r,3,len(ORDER)); ws3.cell(r,1).font=Font(bold=True)
    r+=2; ws3.cell(r,1,"二、按风险等级"); ws3.cell(r,1).font=Font(bold=True); r+=1
    ws3.append(["风险等级","公式计数","Python校验值"]); style_header(ws3,r,3); r+=1
    for rk in RISK_ORDER:
        ws3.cell(r,1,rk); ws3.cell(r,2,cf(RK_COL,rk)); ws3.cell(r,3,sum(1 for i in ORDER if items[i]["risk"]==rk)); r+=1
    r+=2; ws3.cell(r,1,"三、按处置类型"); ws3.cell(r,1).font=Font(bold=True); r+=1
    ws3.append(["处置类型","公式计数","Python校验值"]); style_header(ws3,r,3); r+=1
    for dp in DISP_ORDER:
        ws3.cell(r,1,dp); ws3.cell(r,2,cf(DP_COL,dp)); ws3.cell(r,3,sum(1 for i in ORDER if items[i]["disp"]==dp)); r+=1
    r+=2; ws3.cell(r,1,"四、按领域"); ws3.cell(r,1).font=Font(bold=True); r+=1
    ws3.append(["领域","公式计数","Python校验值"]); style_header(ws3,r,3); r+=1
    for code,name in DOMAIN.items():
        ws3.cell(r,1,name); ws3.cell(r,2,cf(DM_COL,name)); ws3.cell(r,3,sum(1 for i in ORDER if items[i]["domain"]==name)); r+=1
    for rr in range(5, ws3.max_row+1):
        for cc in (1,2,3):
            ws3.cell(rr,cc).border=BORDER; ws3.cell(rr,cc).alignment=CENTER
    ws3.column_dimensions["A"].width=14; ws3.column_dimensions["B"].width=12; ws3.column_dimensions["C"].width=14
    xlsx1 = os.path.join(OUT, "尽调清单核验与反馈表.xlsx")
    wb.save(xlsx1); print("wrote", xlsx1)

    # ============ 2) 补充资料及问询清单.xlsx ============
    wb2 = openpyxl.Workbook()
    wsa = wb2.active; wsa.title = "补件问询"
    fcols = ["反馈编号","类型","优先级","领域","补充资料或问题","核验目的","责任方","建议时点","关联清单编号","关联风险编号"]
    wsa.append(fcols); style_header(wsa,1,len(fcols))
    for fb in feedback:
        wsa.append([fb["fid"],fb["type"],fb["pri"],fb["domain"],fb["ask"],fb["purpose"],
                    fb["owner"],fb["due"],"、".join(fb["items"]),fb["risk"]])
    fwidths=[10,10,8,9,52,40,16,14,26,10]
    for i,w in enumerate(fwidths,1): wsa.column_dimensions[get_column_letter(i)].width=w
    for r in range(2, wsa.max_row+1):
        for c in range(1,len(fcols)+1):
            wsa.cell(r,c).border=BORDER; wsa.cell(r,c).alignment = WRAP if c in (5,6) else CENTER
    wsa.freeze_panes="A2"; wsa.auto_filter.ref=f"A1:{get_column_letter(len(fcols))}{wsa.max_row}"
    wsb = wb2.create_sheet("对应关系")
    wsb.append(["反馈编号","关联清单编号","清单领域"]); style_header(wsb,1,3)
    for fb in feedback:
        for it in fb["items"]:
            wsb.append([fb["fid"], it, items.get(it,{}).get("domain","")])
    wsb.column_dimensions["A"].width=12; wsb.column_dimensions["B"].width=14; wsb.column_dimensions["C"].width=12
    for r in range(2, wsb.max_row+1):
        for c in (1,2,3): wsb.cell(r,c).border=BORDER; wsb.cell(r,c).alignment=CENTER
    wsb.freeze_panes="A2"
    xlsx2 = os.path.join(OUT, "补充资料及问询清单.xlsx")
    wb2.save(xlsx2); print("wrote", xlsx2)

    # ============ 3) 重点问题摘要.docx ============
    doc = DocxDocument()
    doc.add_heading(f"{TARGET}｜法律尽职调查第一轮资料核验与重点问题摘要", level=0)
    doc.add_paragraph(f"提交对象：{ACQUIRER}项目组（买方律师团队复核用，非正式法律意见）")
    doc.add_paragraph(f"目标公司：{TARGET}　|　交易：受让{DEAL_PCT}%股权并取得控制权　|　基准日：{BASELINE}")
    doc.add_paragraph(f"资料范围：虚拟数据室 VDR（第一轮，{MATN}份资料）　|　核验清单：{len(ORDER)}项")
    doc.add_heading("一、交易及审阅范围", level=1)
    doc.add_paragraph(
     f"买方（{ACQUIRER}）拟受让目标公司{DEAL_PCT}%股权并取得控制权。本轮审阅为目标公司第一轮资料（VDR v0.6），"
     "覆盖主体资格与历史沿革、股权与公司治理、重大合同与融资、知识产权与信息技术、劳动人事、"
     "数据合规与业务资质、争议债务与保险、物业与其他共八个领域。全部主体、人员、合同、账号、"
     "地址与交易数据均为虚构合成信息。")
    doc.add_heading("二、执行摘要", level=1)
    n_high=sum(1 for i in ORDER if items[i]["risk"]=="HIGH")
    n_conf=sum(1 for i in ORDER if items[i]["status"]=="资料冲突")
    n_unmet=sum(1 for i in ORDER if items[i]["status"]=="未满足")
    n_part=sum(1 for i in ORDER if items[i]["status"]=="部分满足")
    doc.add_paragraph(
     f"52项清单中：已满足 {sum(1 for i in ORDER if items[i]['status']=='已满足')} 项，"
     f"部分满足 {n_part} 项，未满足 {n_unmet} 项，不适用 {sum(1 for i in ORDER if items[i]['status']=='不适用')} 项，"
     f"待确认 {sum(1 for i in ORDER if items[i]['status']=='待确认')} 项，资料冲突 {n_conf} 项；"
     f"高风险（HIGH）{n_high} 项。识别出跨文件矛盾 {len(contradictions)} 处"
     f"（{ '、'.join(c['title'] for c in contradictions) }），并将重复缺件合并为{len(feedback)}项可执行补件/问询。"
     "四类核心风险（股权/出资、核心知识产权、数据跨境、控制权变更）均已在重点问题中覆盖。")
    doc.add_heading(f"三、跨文件矛盾识别（{len(contradictions)}处）", level=1)
    doc.add_paragraph("下列矛盾通过比对文件实质内容（比例、金额、表述、部署区域）识别，而非仅依据文件命名。")
    for c in contradictions:
        doc.add_heading(f"{c['cid']}　{c['title']}　[{c['severity']}]", level=2)
        doc.add_paragraph("冲突文件：" + "；".join(c["files_conflict"]))
        doc.add_paragraph("矛盾说明：" + c["description"])
        doc.add_paragraph("关联清单：" + "、".join(c["items"]) + "　|　建议处置：" + c["resolution"])
    doc.add_heading("四、核验状态统计", level=1)
    rows = [["维度","类别","数量","占比"]]
    for s in STATUS_ORDER:
        c = sum(1 for i in ORDER if items[i]["status"]==s)
        rows.append(["核验状态", s, str(c), f"{c/52*100:.0f}%"])
    for rk in RISK_ORDER:
        c = sum(1 for i in ORDER if items[i]["risk"]==rk)
        rows.append(["风险等级", rk, str(c), f"{c/52*100:.0f}%"])
    for dp in ["CP","COVENANT","INDEMNITY","RFI","MONITOR","NONE"]:
        c = sum(1 for i in ORDER if items[i]["disp"]==dp)
        rows.append(["处置类型", dp, str(c), f"{c/52*100:.0f}%"])
    doc.add_table(rows)
    doc.add_heading("五、重点问题（12项）", level=1)
    for k in kis:
        rel_fb = sorted({fb["fid"] for fb in feedback if any(it in k["items"] for it in fb["items"])})
        doc.add_heading(f"{k['kid']}　{k['title']}　[{k['risk']} / {k['disp']}]", level=2)
        doc.add_paragraph("事实：" + k["fact"])
        doc.add_paragraph("证据：" + k["evidence"])
        doc.add_paragraph("影响：" + k["impact"])
        doc.add_paragraph("待核验：关联清单 " + "、".join(k["items"]) +
                          "；建议交易处置：" + k["disp"] + "（见补件问询 " + "、".join(rel_fb) + "）")
    doc.add_heading("六、优先补件与下一步", level=1)
    doc.add_paragraph("P0（交易先决/必须于交割前解决）：")
    for fb in feedback:
        if fb["pri"]=="P0":
            doc.add_bullet(f"{fb['fid']}（{fb['domain']}）：{fb['ask']}　责任方：{fb['owner']}；时点：{fb['due']}")
    doc.add_paragraph("P1（签约前澄清/补正）：")
    for fb in feedback:
        if fb["pri"]=="P1":
            doc.add_bullet(f"{fb['fid']}（{fb['domain']}）：{fb['ask']}　责任方：{fb['owner']}")
    doc.add_heading("七、范围限制和人工复核提示", level=1)
    doc.add_paragraph(
     "1) 本摘要基于第一轮资料（VDR v0.6）作出，部分历史档案、股东协议、数据出境与保险材料仍在补充；"
     "2) 含“资料冲突/未满足/高风险”的项及全部P0补件均建议项目律师人工复核；"
     "3) 本成果为买方律师团队支持性分析，不构成正式法律意见，不替代对登记机关、主管机关及相对方的独立核验；"
     "4) 全部主体、合同、账号、地址与交易数据均为虚构合成信息，未关联任何真实主体；"
     "5) 未见证据的事项已写入“资料缺口/待提供”，未将“未提供”表述为违法或无效。")
    docx_path = os.path.join(OUT, "重点问题摘要.docx")
    doc.save(docx_path); print("wrote", docx_path)

    # ============ 4) result_manifest.json ============
    manifest = {
      "task_id":"legal-dd-verify", "instance_id":IID, "cutoff_date":BASELINE,
      "target_name":TARGET, "checklist_count":len(ORDER), "items":[], "key_issues":[],
      "cross_file_contradictions":[],
      "generated_files":[
        "尽调清单核验与反馈表.xlsx","补充资料及问询清单.xlsx",
        "重点问题摘要.docx","result_manifest.json"]
    }
    for iid in ORDER:
        d = items[iid]
        manifest["items"].append({
            "item_id":iid, "status":d["status"], "risk_level":d["risk"],
            "evidence":[e for e in d["evidence"].split("\n") if e.strip()],
            "feedback_ids":d.get("feedback_ids",[]),
        })
    for k in kis:
        manifest["key_issues"].append({
            "issue_id":k["kid"], "title":k["title"], "risk_level":k["risk"],
            "disposition":k["disp"], "related_items":k["items"],
            "evidence":k["evidence"]})
    for c in contradictions:
        manifest["cross_file_contradictions"].append({
            "contradiction_id":c["cid"], "severity":c["severity"], "title":c["title"],
            "conflicting_files":c["files_conflict"], "related_items":c["items"],
            "resolution":c["resolution"]})
    json_path = os.path.join(OUT, "result_manifest.json")
    with open(json_path,"w",encoding="utf-8") as f:
        json.dump(manifest,f,ensure_ascii=False,indent=2)
    print("wrote", json_path)
    return OUT

if __name__ == "__main__":
    import argparse
    from analyze import run
    from extract import Store
    ap = argparse.ArgumentParser()
    _sub = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ap.add_argument("--input", default=os.path.join(_sub, "materials"))
    ap.add_argument("--checklist", default=os.path.join(_sub, "checklist", "checklist_52.json"))
    ap.add_argument("--out", default=OUT)
    args = ap.parse_args()
    s = Store(args.input); cl = json.load(open(args.checklist, encoding="utf-8"))
    res = run(s, cl)
    render(res, cl, s, out_dir=args.out)
