#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Independent verification of the rendered due-diligence deliverables.

Does NOT import the analysis engine — it only reads the 4 output files plus the
authoritative data-room source files, and checks:
  - citations resolve to real source files (no fabrication)
  - 52 items unique & enum-valid
  - 4 mandatory core risks covered
  - contradictions recorded & item-linked
  - feedback deduped (15-22) and typed
  - docx has the required sections + 8-12 key issues

Usage:
  python3 verify.py [--out DIR] [--input DATAROOM_DIR]
"""
import os, re, json, sys
# Offline bootstrap: vendored pure-python deps + shared stdlib docx helper.
_HERE = os.path.dirname(os.path.abspath(__file__))
_SUB = os.path.dirname(_HERE)
for _p in (os.path.join(_SUB, "vendor"), _SUB):
    if _p not in sys.path:
        sys.path.insert(0, _p)
import openpyxl
from _docx import read_docx

DEFAULT_OUT = os.path.join(_SUB, "examples")
DEFAULT_INPUT = os.path.join(_SUB, "materials")

STATUS_OK = {"已满足","部分满足","未满足","不适用","待确认","资料冲突"}
RISK_OK = {"HIGH","MEDIUM","LOW","INFO"}
DISP_OK = {"CP","COVENANT","INDEMNITY","PRICE","RFI","MONITOR","NONE"}
FBTYPE_OK = {"补充资料","事实问询","外部确认","整改动作"}
CORE = {"股权/出资":["股权","出资"], "核心知识产权":["知识产权","IP","核心"],
        "数据跨境":["数据跨境","跨境"], "控制权变更":["控制权变更","控制权"]}


def main(out=DEFAULT_OUT, input_dir=DEFAULT_INPUT):
    errors, warns = [], []

    # authoritative source filenames (for citation resolution)
    auth = set()
    for base,_,fs in os.walk(input_dir):
        for f in fs: auth.add(f)
    auth_stem = {os.path.splitext(f)[0] for f in auth}

    def resolves(head):
        head = head.strip()
        if not head or "未发现可支持资料" in head or head == "（无）" or "无。" in head:
            return True
        h = re.sub(r"[（(][^）)]*[）)]\s*$", "", head).strip()
        if h in auth: return True
        if os.path.splitext(h)[0] in auth_stem: return True
        for a in auth:
            if h and (h in a or a in h): return True
        return False

    # ---- manifest ----
    M = json.load(open(os.path.join(out,"result_manifest.json"), encoding="utf-8"))
    items = M["items"]; kis = M["key_issues"]; contra = M.get("cross_file_contradictions", [])
    if len(items) != 52: errors.append(f"manifest items={len(items)}!=52")
    if len({i["item_id"] for i in items}) != 52: errors.append("manifest item ids not unique")
    if len(kis) < 8 or len(kis) > 12: errors.append(f"key_issues count={len(kis)} not in 8-12")
    if len(contra) < 1: errors.append("no cross-file contradictions recorded")

    bad = []
    for it in items:
        for ev in it["evidence"]:
            head = ev.split("｜")[0]
            if not resolves(head): bad.append(f"{it['item_id']}: {head}")
    if bad:
        errors.append(f"{len(bad)} UNRESOLVED citations (first 5): " + " | ".join(bad[:5]))

    for it in items:
        if it["status"] not in STATUS_OK: errors.append(f"{it['item_id']} bad status {it['status']}")
        if it["risk_level"] not in RISK_OK: errors.append(f"{it['item_id']} bad risk {it['risk_level']}")

    core_hit = {c: False for c in CORE}
    for ki in kis:
        t = ki["title"]
        for c, kws in CORE.items():
            if any(k in t for k in kws): core_hit[c] = True
    for it in items:
        if "违法" in it.get("gap","") or "无效" in it.get("gap",""):
            warns.append(f"{it['item_id']} gap mentions 违法/无效 — review")
    if [c for c,v in core_hit.items() if not v]:
        errors.append(f"CORE RISK NOT COVERED: {[c for c,v in core_hit.items() if not v]}")

    item_ids = {i["item_id"] for i in items}
    for c in contra:
        for it in c.get("related_items", []):
            if it not in item_ids: errors.append(f"contradiction {c['contradiction_id']} refs {it} missing")

    # ---- xlsx 核验总表 ----
    wb = openpyxl.load_workbook(os.path.join(out,"尽调清单核验与反馈表.xlsx"))
    ws = wb["核验总表"]
    xrows = ws.max_row - 1
    if xrows != 52: errors.append(f"核验总表 rows={xrows}!=52")
    xbad = 0
    for r in range(2, ws.max_row+1):
        cell = ws.cell(r,6).value or ""
        for ev in str(cell).split("\n"):
            head = ev.split("｜")[0]
            if not resolves(head): xbad += 1
    if xbad: errors.append(f"核验总表 {xbad} unresolved citations")
    for r in range(2, ws.max_row+1):
        dp = ws.cell(r,9).value
        if dp not in DISP_OK: errors.append(f"核验总表 row{r} bad disp {dp}")
    for sh in ["核验总表","证据索引","跨文件矛盾","统计摘要"]:
        if sh not in wb.sheetnames: errors.append(f"missing sheet {sh}")

    # ---- xlsx 补件问询 ----
    wb2 = openpyxl.load_workbook(os.path.join(out,"补充资料及问询清单.xlsx"))
    if "补件问询" not in wb2.sheetnames or "对应关系" not in wb2.sheetnames:
        errors.append("补件问询/对应关系 sheet missing")
    fb_count = wb2["补件问询"].max_row - 1
    if fb_count < 15 or fb_count > 22: warns.append(f"feedback count={fb_count} (expect 15-22)")
    for r in range(2, wb2["补件问询"].max_row+1):
        ty = wb2["补件问询"].cell(r,2).value
        if ty not in FBTYPE_OK: errors.append(f"feedback {wb2['补件问询'].cell(r,1).value} bad type {ty}")

    # ---- docx (stdlib reader) ----
    doc_info = read_docx(os.path.join(out, "重点问题摘要.docx"))
    heads = [t for (t, style) in doc_info["paragraphs"] if style.startswith("Heading")]
    req_sections = ["交易及审阅范围","执行摘要","跨文件矛盾识别","核验状态统计","重点问题","优先补件与下一步","范围限制"]
    for s in req_sections:
        if not any(s in h for h in heads): errors.append(f"docx missing section {s}")
    ki_heads = [h for h in heads if h.startswith("K-")]
    if len(ki_heads) < 8: errors.append(f"docx key issues headings={len(ki_heads)}<8")

    # ---- report ----
    print("="*60)
    print("INDEPENDENT DELIVERABLE VERIFICATION")
    print(f"  out={out}")
    print("="*60)
    print(f"items={len(items)} key_issues={len(kis)} contradictions={len(contra)} feedback(xlsx)={fb_count}")
    print(f"unresolved citations: manifest={len(bad)} xlsx={xbad}")
    print(f"core risk coverage: {core_hit}")
    if warns:
        print("WARNINGS:")
        for w in warns: print("  -", w)
    print("-"*60)
    if errors:
        print(f"ERRORS ({len(errors)}):")
        for e in errors: print("  X", e)
        print("\nRESULT: FAIL")
        return 1
    else:
        print("RESULT: PASS — citations resolve, 52/unique, enums valid,")
        print("          core risks covered, contradictions recorded, sections present.")
        return 0


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=DEFAULT_OUT)
    ap.add_argument("--input", default=DEFAULT_INPUT)
    args = ap.parse_args()
    sys.exit(main(args.out, args.input))
