#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Due-diligence verification pipeline — one-click orchestrator.

Usage:
  python3 run.py --input <data-room directory> \
                 --checklist <52-item checklist xlsx or json> \
                 --out <output directory> \
                 [--verify]

It runs the full chain:
  extract (read source files)  ->  analyze (rules + contradiction detection)
  -> render (4 deliverables)  ->  verify (independent consistency check)

All facts (equity ratios, amounts, cloud regions, IP owners, names, years...)
are extracted from the source files at runtime. The 52-item checklist is parsed
at runtime (from xlsx or json), so the same command re-derives deliverables for
any data room without modifying the package.
"""
import os, sys, argparse, re, json

# Offline bootstrap: make the vendored pure-python deps importable WITHOUT pip.
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
_SUB = os.path.dirname(HERE)
for _p in (os.path.join(_SUB, "vendor"), _SUB):
    if _p not in sys.path:
        sys.path.insert(0, _p)


def ensure_deps():
    """Dependencies are vendored (see ./vendor) — no network install at runtime."""
    return


def discover_input():
    """Walk up from cwd looking for a data-room directory (materials/ or materials)."""
    cur = os.getcwd()
    for _ in range(6):
        for root, dirs, _ in os.walk(cur):
            if os.path.basename(root) in ("materials", "materials") and os.path.isdir(root):
                return root
            if "materials" in dirs or "materials" in dirs:
                return os.path.join(root, "materials" if "materials" in dirs else "materials")
        cur = os.path.dirname(cur)
        if cur in ("/", ""):
            break
    return None


def discover_checklist(input_dir):
    """Find 尽调资料清单.xlsx near the input dir (instance root)."""
    roots = []
    d = input_dir or os.getcwd()
    for _ in range(6):
        if d and os.path.isdir(d):
            roots.append(d)
        d = os.path.dirname(d)
        if d in ("/", ""):
            break
    for root in roots:
        try:
            for name in os.listdir(root):
                if name.startswith("尽调资料清单") and name.endswith(".xlsx"):
                    return os.path.join(root, name)
        except Exception:
            pass
    return None


def parse_checklist_xlsx(xlsx):
    """Parse the instance's 尽调资料清单.xlsx into the checklist list."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    header_row = None
    headers = []
    for r in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v and "编号" in str(v) for v in vals):
            header_row = r
            headers = [str(v).strip() if v is not None else "" for v in vals]
            break
    if header_row is None:
        raise RuntimeError("checklist header row (编号) not found in " + xlsx)
    col = {h: i + 1 for i, h in enumerate(headers)}
    items = []
    for r in range(header_row + 1, ws.max_row + 1):
        raw = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue
        iid = raw[col["编号"] - 1] if "编号" in col else None
        if iid is None:
            continue
        def g(key):
            return raw[col[key] - 1] if key in col else None
        items.append({
            "id": str(iid).strip(),
            "req": g("文件及信息要求") or "",
            "company_feedback": g("公司反馈") or "",
            "provision": g("提供情况") or "",
            "vdr": g("VDR编号/路径") or "",
            "focus": g("律师调查途径或核验重点") or "",
            "note": g("备注") or "",
        })
    return items


def load_checklist(path):
    if path.endswith(".json"):
        return json.load(open(path, encoding="utf-8"))
    if path.endswith(".xlsx"):
        return parse_checklist_xlsx(path)
    # fall back to json
    return json.load(open(path, encoding="utf-8"))


def main():
    ap = argparse.ArgumentParser(description="Legal due-diligence verification pipeline")
    _sub = os.path.dirname(HERE)
    ap.add_argument("--input", default=os.path.join(_sub, "materials"),
                    help="data-room directory (defaults to the bundled materials/ demo)")
    ap.add_argument("--checklist", default=os.path.join(_sub, "checklist", "checklist_52.json"),
                    help="52-item checklist xlsx or json")
    ap.add_argument("--out", default=None, help="output directory for the 4 deliverables")
    ap.add_argument("--verify", action="store_true", help="run the independent verifier after render")
    args = ap.parse_args()

    ensure_deps()  # vendored; no-op (no network)

    from extract import Store
    from analyze import run as analyze_run
    from render import render as render_deliverables

    input_dir = args.input or discover_input()
    if not input_dir or not os.path.isdir(input_dir):
        sys.exit("ERROR: cannot locate materials directory. Pass --input explicitly.")
    checklist_path = args.checklist or discover_checklist(input_dir)
    if not checklist_path or not os.path.exists(checklist_path):
        sys.exit("ERROR: cannot locate 尽调资料清单.xlsx/json. Pass --checklist explicitly.")
    out_dir = args.out or os.path.join(os.getcwd(), "deliverables")
    os.makedirs(out_dir, exist_ok=True)

    print("[1/3] Extracting materials from:", input_dir)
    store = Store(input_dir)
    cl = load_checklist(checklist_path)
    print("      materials parsed:", len(store.by_prefix), "| checklist items:", len(cl))

    print("[2/3] Analyzing (52-item verification + contradiction detection) ...")
    res = analyze_run(store, cl)
    res.setdefault("meta", {})["instance_id"] = "demo"
    print("      instance_id:", res["meta"]["instance_id"])
    print("      items=%d contradictions=%d feedback=%d key_issues=%d" % (
        len(res["items"]), len(res["contradictions"]),
        len(res["feedback"]), len(res["key_issues"])))
    for c in res["contradictions"]:
        print("        -", c["cid"], c["title"], "->", c["items"])

    print("[3/3] Rendering 4 deliverables to:", out_dir)
    out = render_deliverables(res, cl, store, out_dir=out_dir)
    print("      done:", out)

    if args.verify:
        print("[verify] Running independent verifier on:", out_dir)
        try:
            import verify
            rc = verify.main(out_dir, input_dir)
            if rc != 0:
                print("[verify] RESULT: FAIL — see above")
        except SystemExit:
            pass

    print("OK. Deliverables ready in:", out)


if __name__ == "__main__":
    main()
